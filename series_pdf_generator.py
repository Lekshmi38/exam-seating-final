"""
series_pdf_generator.py
-----------------------
Generates two output files for a Series Exam arrangement:

  1. qp_counts.xlsx          – QP count per subject (summary + room-wise)
  2. seating_arrangement.pdf – room-wise seating chart (landscape, one room per section)
  3. master_list.pdf         – CLASS | ROLL RANGE | HALL NO

The class_assignment.xlsx has been replaced by qp_counts.xlsx.
Student tokens in the PDF now show the FULL class name (e.g. S7CS1, not CS1).
"""

import os
import re
from collections import defaultdict

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)

# ---------------------------------------------------------------------------
# Hall number mapping
# ---------------------------------------------------------------------------
HALL_NUMBERS = {
    "Room 1":  "108", "Room 2":  "109", "Room 3":  "111", "Room 4":  "112",
    "Room 5":  "208", "Room 6":  "209", "Room 7":  "211", "Room 8":  "212",
    "Room 9":  "308", "Room 10": "309", "Room 11": "311", "Room 12": "312",
    "Room 13": "313", "Room 14": "314", "Room 15": "315", "Room 16": "316",
    "Room 17": "317", "Room 18": "318", "Room 19": "319", "Room 20": "320",
    "Room 21": "321", "Room 22": "322", "Room 23": "323", "Room 24": "324",
    "Room 25": "325", "Room 26": "326", "Room 27": "327", "Room 28": "328",
}

INSTITUTION = "LBS INSTITUTE OF TECHNOLOGY FOR WOMEN, POOJAPPURA"


def _hall(room_name):
    if room_name in HALL_NUMBERS:
        return HALL_NUMBERS[room_name]
    m = re.search(r'(\d+)', room_name)
    return str(300 + int(m.group(1))) if m else room_name


def _room_sort_key(name):
    m = re.search(r'\d+', name)
    return int(m.group()) if m else 0


def _parse_token(token):
    """
    Parse a student token like 'S7CS1_3' into (class_name, roll_no).
    Returns ('--', '') for empty/dash tokens.
    Token format: <ClassName>_<RollNumber>
    e.g.  S7CS1_3   → ('S7CS1', '3')
          S5EC_12   → ('S5EC',  '12')
          --        → ('--',    '')
    """
    if not token or token == '--':
        return ('--', '')
    parts = str(token).rsplit('_', 1)
    if len(parts) == 2:
        return (parts[0], parts[1])
    return (str(token), '')


# ===========================================================================
# 1.  QP COUNT EXCEL  (replaces class_assignment.xlsx)
# ===========================================================================

def create_series_qp_excel(arrangement, output_path):
    """
    Create qp_counts.xlsx with:
      Sheet 1 "Subject Summary"  – Subject | Total QP Count
      Sheet 2 "Room-wise Counts" – Room | Hall No | Subject | Count
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rooms_data = arrangement.get('rooms', {})
        if not rooms_data:
            print("WARNING: No rooms data; skipping QP Excel generation")
            return False

        # ── build counts from student tokens ─────────────────────────────────
        # subject_totals:  { subject_name: total_count }
        # room_subject_counts: { room_name: { subject_name: count } }
        subject_totals = defaultdict(int)
        room_subject_counts = {}

        sorted_rooms = sorted(rooms_data.keys(), key=_room_sort_key)

        for room_name in sorted_rooms:
            room_data = rooms_data[room_name]
            room_subject_counts[room_name] = defaultdict(int)

            # Build class → subject mapping from subjects list
            # subjects list entries look like "S7CS1: Computer Networks"
            class_subject_map = {}
            for subj_str in room_data.get('subjects', []):
                if ':' in subj_str:
                    cls_part, subj_part = subj_str.split(':', 1)
                    class_subject_map[cls_part.strip()] = subj_part.strip()

            for block_name, block_data in room_data.get('blocks', {}).items():
                for token in block_data.get('students', []):
                    cls_name, roll = _parse_token(token)
                    if cls_name == '--':
                        continue
                    subj = class_subject_map.get(cls_name, 'Unknown')
                    room_subject_counts[room_name][subj] += 1
                    subject_totals[subj] += 1

        # ── openpyxl styles ───────────────────────────────────────────────────
        HEADER_FILL = PatternFill("solid", start_color="1A5276", end_color="1A5276")
        ALT_FILL    = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
        TOTAL_FILL  = PatternFill("solid", start_color="D5DBDB", end_color="D5DBDB")
        WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        B_FONT  = Font(name="Arial", size=10)
        BLD     = Font(name="Arial", bold=True, size=10)
        CENTER  = Alignment(horizontal="center", vertical="center")
        LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin    = Side(style="thin",   color="AEB6BF")
        thick   = Side(style="medium", color="1A5276")
        TB      = Border(left=thin, right=thin, top=thin, bottom=thin)
        HB      = Border(left=thick, right=thick, top=thick, bottom=thick)

        wb = Workbook()
        wb.remove(wb.active)

        # ── Sheet 1: Subject Summary ──────────────────────────────────────────
        ws1 = wb.create_sheet("Subject Summary")

        ws1.merge_cells("A1:C1")
        ws1["A1"] = f"{INSTITUTION} — Series Exam QP Count Summary"
        ws1["A1"].font = Font(name="Arial", bold=True, size=12, color="1A5276")
        ws1["A1"].alignment = CENTER

        for col, hdr in enumerate(["Sl.No", "Subject", "QP Count Required"], 1):
            c = ws1.cell(row=2, column=col, value=hdr)
            c.font = H_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = HB

        grand_total = 0
        for idx, (subj, cnt) in enumerate(sorted(subject_totals.items()), 1):
            fill = ALT_FILL if idx % 2 == 0 else WHITE_FILL
            row_n = idx + 2
            for col, val in enumerate([idx, subj, cnt], 1):
                c = ws1.cell(row=row_n, column=col, value=val)
                c.font = B_FONT; c.fill = fill; c.border = TB
                c.alignment = LEFT if col == 2 else CENTER
            grand_total += cnt

        tot_r = len(subject_totals) + 3
        ws1.merge_cells(f"A{tot_r}:B{tot_r}")
        ws1[f"A{tot_r}"] = "GRAND TOTAL"
        ws1[f"A{tot_r}"].font = BLD; ws1[f"A{tot_r}"].fill = TOTAL_FILL
        ws1[f"A{tot_r}"].alignment = CENTER
        ws1[f"C{tot_r}"] = grand_total
        ws1[f"C{tot_r}"].font = BLD; ws1[f"C{tot_r}"].fill = TOTAL_FILL
        ws1[f"C{tot_r}"].alignment = CENTER

        for col, w in zip("ABC", [8, 40, 18]):
            ws1.column_dimensions[get_column_letter({"A":1,"B":2,"C":3}[col])].width = w
        ws1.column_dimensions["A"].width = 8
        ws1.column_dimensions["B"].width = 40
        ws1.column_dimensions["C"].width = 18

        # ── Sheet 2: Room-wise Counts ─────────────────────────────────────────
        ws2 = wb.create_sheet("Room-wise Counts")

        ws2.merge_cells("A1:E1")
        ws2["A1"] = f"{INSTITUTION} — Room-wise QP Distribution"
        ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="1A5276")
        ws2["A1"].alignment = CENTER

        for col, hdr in enumerate(["Room", "Hall No", "Subject", "QP Count", "% of Total"], 1):
            c = ws2.cell(row=2, column=col, value=hdr)
            c.font = H_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = HB

        row_n = 3
        for idx_r, room_name in enumerate(sorted_rooms):
            subj_counts = room_subject_counts.get(room_name, {})
            if not subj_counts:
                continue
            hall = _hall(room_name)
            fill = ALT_FILL if idx_r % 2 == 0 else WHITE_FILL

            for subj, cnt in sorted(subj_counts.items()):
                pct = round(cnt / subject_totals[subj] * 100, 1) if subject_totals[subj] else 0
                for col, val in enumerate([room_name, hall, subj, cnt, f"{pct}%"], 1):
                    c = ws2.cell(row=row_n, column=col, value=val)
                    c.font = B_FONT; c.fill = fill; c.border = TB
                    c.alignment = LEFT if col == 3 else CENTER
                row_n += 1

        # totals row
        ws2.merge_cells(f"A{row_n}:C{row_n}")
        ws2[f"A{row_n}"] = "TOTAL"
        ws2[f"A{row_n}"].font = BLD; ws2[f"A{row_n}"].fill = TOTAL_FILL
        ws2[f"A{row_n}"].alignment = CENTER
        ws2[f"D{row_n}"] = grand_total
        ws2[f"D{row_n}"].font = BLD; ws2[f"D{row_n}"].fill = TOTAL_FILL
        ws2[f"D{row_n}"].alignment = CENTER

        for col_letter, w in zip("ABCDE", [12, 10, 38, 12, 12]):
            ws2.column_dimensions[col_letter].width = w

        wb.save(output_path)
        print(f"✓ QP count Excel saved: {output_path}")
        return True

    except Exception as e:
        import traceback
        print(f"Error creating QP count Excel: {e}")
        traceback.print_exc()
        return False


# ===========================================================================
# 2.  SEATING ARRANGEMENT PDF  (room-wise, landscape A4)
#     FIX: show FULL class name (S7CS1) + roll number clearly, not stripped
# ===========================================================================

def create_series_seating_pdf(arrangement, output_path, exam_info=None):
    """
    Landscape PDF showing each room's block-wise seating arrangement.
    Student cells show:  S7CS1          ← full class name, bold
                         #3             ← roll number
    """
    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(A4),
            rightMargin=0.5 * cm, leftMargin=0.5 * cm,
            topMargin=1.0 * cm,   bottomMargin=1.0 * cm,
        )

        BLOCK_ORDER = ["Left1", "Left2", "Left3", "Middle1", "Middle2",
                       "Right1", "Right2", "Right3"]

        def _ps(name, size, bold=False, align=TA_CENTER, color=colors.black, sa=2):
            return ParagraphStyle(
                name,
                fontName="Helvetica-Bold" if bold else "Helvetica",
                fontSize=size, textColor=color,
                alignment=align, spaceAfter=sa, leading=size + 2,
            )

        title_style   = _ps("ST",  13, bold=True)
        sub_style     = _ps("SS",   9, bold=False)
        room_style    = _ps("SR",  10, bold=True, color=colors.HexColor("#1A5276"))
        col_hdr_style = _ps("SCH",  8, bold=True, color=colors.white)
        row_num_style = _ps("RN",   7, bold=True,
                            color=colors.HexColor("#555555"), align=TA_CENTER)

        # Two-line student cell: class name bold on top, roll# smaller below
        cls_style  = _ps("CLS",  7, bold=True,  align=TA_CENTER,
                         color=colors.HexColor("#1A5276"))
        roll_style = _ps("ROLL", 6, bold=False, align=TA_CENTER,
                         color=colors.HexColor("#555555"))
        dash_style = _ps("DASH", 8, bold=False, align=TA_CENTER,
                         color=colors.HexColor("#AAAAAA"))

        elements = []

        # ── document header ───────────────────────────────────────────────────
        elements.append(Paragraph(INSTITUTION, title_style))
        if exam_info:
            if exam_info.get('title'):
                elements.append(Paragraph(exam_info['title'], sub_style))
            if exam_info.get('date'):
                elements.append(Paragraph(f"Date: {exam_info['date']}", sub_style))
        elements.append(Paragraph("SERIES EXAM — ROOM-WISE SEATING ARRANGEMENT", sub_style))
        elements.append(Spacer(1, 0.4 * cm))

        rooms_data   = arrangement.get('rooms', {})
        sorted_rooms = sorted(rooms_data.keys(), key=_room_sort_key)

        for room_idx, room_name in enumerate(sorted_rooms):
            room_data = rooms_data[room_name]
            hall      = _hall(room_name)

            elements.append(Paragraph(
                f"{room_name}  (Hall No: {hall})  —  Total: {room_data['total']} / {room_data['capacity']} Students",
                room_style))

            # subjects line  – show class: subject mapping compactly
            subjects_text = "  |  ".join(room_data.get('subjects', []))
            elements.append(Paragraph(
                subjects_text,
                _ps("SSubj", 7, color=colors.HexColor("#555555"), align=TA_LEFT)
            ))
            elements.append(Spacer(1, 0.1 * cm))

            blocks = room_data.get('blocks', {})
            present_blocks = [b for b in BLOCK_ORDER
                              if b in blocks and blocks[b]['capacity'] > 0]

            if not present_blocks:
                elements.append(Spacer(1, 0.3 * cm))
                continue

            max_rows = max(blocks[b]['capacity'] for b in present_blocks)

            # ── header row ────────────────────────────────────────────────────
            hdr = [Paragraph("Row", col_hdr_style)]
            for b in present_blocks:
                cap = blocks[b]['capacity']
                cnt = blocks[b]['count']
                hdr.append(Paragraph(f"{b}\n{cnt}/{cap}", col_hdr_style))

            table_data = [hdr]

            # ── data rows ─────────────────────────────────────────────────────
            for r in range(max_rows):
                row = [Paragraph(str(r + 1), row_num_style)]
                for b in present_blocks:
                    students = blocks[b].get('students', [])
                    token    = students[r] if r < len(students) else '--'
                    cls_name, roll_no = _parse_token(token)

                    if cls_name == '--':
                        cell_content = Paragraph("—", dash_style)
                    else:
                        # Two-line cell: full class name + roll number
                        # Use a mini-table inside the cell for clean two-line layout
                        from reportlab.platypus import KeepTogether
                        cell_content = Table(
                            [
                                [Paragraph(cls_name, cls_style)],
                                [Paragraph(f"#{roll_no}", roll_style)],
                            ],
                            colWidths=None,
                            style=[
                                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                                ('TOPPADDING',    (0, 0), (-1, -1), 1),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                                ('LEFTPADDING',   (0, 0), (-1, -1), 1),
                                ('RIGHTPADDING',  (0, 0), (-1, -1), 1),
                            ]
                        )
                    row.append(cell_content)
                table_data.append(row)

            # ── column widths ─────────────────────────────────────────────────
            usable_w   = 27.0 * cm   # landscape A4 minus margins ≈ 28cm, keep margin
            row_col_w  = 0.9 * cm
            data_col_w = (usable_w - row_col_w) / len(present_blocks)
            col_widths = [row_col_w] + [data_col_w] * len(present_blocks)

            # ── row heights: header taller, data rows accommodate 2 lines ─────
            row_heights = [0.9 * cm] + [1.1 * cm] * max_rows

            tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights,
                        repeatRows=1)
            tbl_style = [
                # header
                ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor("#1A5276")),
                ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
                ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, 0),  8),
                # all cells
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor("#AEB6BF")),
                ('BOX',           (0, 0), (-1, -1), 1.2, colors.HexColor("#1A5276")),
                ('TOPPADDING',    (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING',   (0, 0), (-1, -1), 2),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
                # row-number column
                ('BACKGROUND',    (0, 1), (0, -1),  colors.HexColor("#F2F3F4")),
                ('FONTNAME',      (0, 1), (0, -1),  'Helvetica-Bold'),
                ('FONTSIZE',      (0, 1), (0, -1),  7),
            ]
            # alternate row shading
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    tbl_style.append(
                        ('BACKGROUND', (1, i), (-1, i), colors.HexColor("#EBF5FB"))
                    )
            tbl.setStyle(TableStyle(tbl_style))
            elements.append(tbl)
            elements.append(Spacer(1, 0.5 * cm))

            # page break every 2 rooms (keep pairs together on same page)
            if (room_idx + 1) % 2 == 0 and room_idx < len(sorted_rooms) - 1:
                elements.append(PageBreak())

        doc.build(elements)
        print(f"✓ Series seating PDF saved: {output_path}")
        return True

    except Exception as e:
        import traceback
        print(f"Error creating series seating PDF: {e}")
        traceback.print_exc()
        return False


# ===========================================================================
# 3.  MASTER LIST PDF  (CLASS | ROLL RANGE | HALL NO)
# ===========================================================================

def create_series_master_pdf(arrangement, output_path, exam_info=None):
    """
    Portrait A4 master list:
      CLASS  |  ROLL NUMBERS (1 to N)  |  HALL NO
    """
    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=1.5 * cm, leftMargin=1.5 * cm,
            topMargin=1.5 * cm,   bottomMargin=1.5 * cm,
        )

        def _ps(name, size, bold=False, align=TA_CENTER, color=colors.black, sa=4):
            return ParagraphStyle(
                name,
                fontName="Helvetica-Bold" if bold else "Helvetica",
                fontSize=size, textColor=color,
                alignment=align, spaceAfter=sa,
            )

        hdr1   = _ps("MH1", 13, bold=True, sa=3)
        hdr2   = _ps("MH2", 11, bold=True, sa=3)
        hdr3   = _ps("MH3", 10, sa=10)
        col_h  = _ps("MCH", 10, bold=True, color=colors.white)
        cell_c = _ps("MCC",  9, bold=True)
        cell_l = _ps("MCL",  9, align=TA_LEFT)

        elements = []
        elements.append(Paragraph(INSTITUTION, hdr1))
        elements.append(Paragraph(
            "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY (APJAKTU)", hdr2))
        if exam_info:
            if exam_info.get('title'):
                elements.append(Paragraph(exam_info['title'], hdr2))
            if exam_info.get('date'):
                elements.append(Paragraph(f"Series Exam — {exam_info['date']}", hdr3))
        else:
            elements.append(Spacer(1, 0.3 * cm))

        # ── build data from classwise_table ───────────────────────────────────
        classwise = arrangement.get('classwise_table', {})

        class_hall_rows = {}
        for cls, entries in classwise.items():
            hall_chunks = []
            for entry in entries:
                hall = _hall(entry.get("Room", ""))
                rs   = entry.get("Roll Start", "")
                re_  = entry.get("Roll End",   "")
                hall_chunks.append([hall, rs, re_])
            # merge consecutive entries in same hall
            merged = []
            for hall, rs, re_ in hall_chunks:
                if (merged and merged[-1][0] == hall
                        and isinstance(merged[-1][2], int)
                        and isinstance(rs, int)
                        and rs == merged[-1][2] + 1):
                    merged[-1][2] = re_
                else:
                    merged.append([hall, rs, re_])
            class_hall_rows[cls] = merged

        COL_W = [3 * cm, 10 * cm, 3 * cm]

        table_data = [[
            Paragraph("CLASS",        col_h),
            Paragraph("ROLL NUMBERS", col_h),
            Paragraph("HALL NO",      col_h),
        ]]

        ALT_A = colors.HexColor("#EBF5FB")
        ALT_B = colors.white
        row_fills = []

        for idx, cls in enumerate(sorted(class_hall_rows.keys())):
            rows = class_hall_rows[cls]
            fill = ALT_A if idx % 2 == 0 else ALT_B

            for i, (hall, rs, re_) in enumerate(rows):
                if isinstance(rs, int) and isinstance(re_, int):
                    roll_str = f"Roll {rs} to {re_}  ({re_ - rs + 1} students)"
                else:
                    roll_str = f"Roll {rs} to {re_}"

                cls_para  = Paragraph(cls,      cell_c) if i == 0 else Paragraph("", cell_c)
                roll_para = Paragraph(roll_str, cell_l)
                hall_para = Paragraph(str(hall), cell_c)
                table_data.append([cls_para, roll_para, hall_para])
                row_fills.append((len(table_data) - 1, fill))

        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1A5276")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  10),
            ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#AEB6BF")),
            ("BOX",           (0, 0), (-1, -1), 1.2, colors.HexColor("#1A5276")),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.5, colors.HexColor("#1A5276")),
        ]
        for ridx, fill in row_fills:
            style_cmds.append(("BACKGROUND", (0, ridx), (-1, ridx), fill))

        tbl = Table(table_data, colWidths=COL_W, repeatRows=1)
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)

        doc.build(elements)
        print(f"✓ Series master list PDF saved: {output_path}")
        return True

    except Exception as e:
        import traceback
        print(f"Error creating series master PDF: {e}")
        traceback.print_exc()
        return False


# ===========================================================================
# 4.  CONVENIENCE WRAPPER
# ===========================================================================

def save_series_files(arrangement, output_folder, exam_info=None):
    """
    Save all three series-exam output files into `output_folder`.
    Returns dict of {'qp_excel': path, 'seating_pdf': path, 'master_pdf': path}

    Note: class_assignment.xlsx has been replaced by qp_counts.xlsx
    """
    os.makedirs(output_folder, exist_ok=True)

    qp_excel_path    = os.path.join(output_folder, "qp_counts.xlsx")
    seating_pdf_path = os.path.join(output_folder, "seating_arrangement.pdf")
    master_pdf_path  = os.path.join(output_folder, "master_list.pdf")

    results = {}

    if create_series_qp_excel(arrangement, qp_excel_path):
        results['qp_excel'] = qp_excel_path

    if create_series_seating_pdf(arrangement, seating_pdf_path, exam_info=exam_info):
        results['seating_pdf'] = seating_pdf_path

    if create_series_master_pdf(arrangement, master_pdf_path, exam_info=exam_info):
        results['master_pdf'] = master_pdf_path

    return results
