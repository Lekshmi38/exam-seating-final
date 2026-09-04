"""
Seating Arrangement TXT -> Excel Converter (to be used by Flask app)
=============================================================================
This module converts seating report TXT files to properly formatted Excel files
with metadata extracted from the TXT content.
"""

import re
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# =============================================================================
# Style helpers
# =============================================================================

def S(v):
    return Side(style=v) if v else Side(style=None)


def brd(L=None, R=None, T=None, B=None):
    return Border(left=S(L), right=S(R), top=S(T), bottom=S(B))


def fnt(name="Cambria", size=12, bold=False):
    return Font(name=name, size=size, bold=bold)


def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# Re-usable border constants
B_ALL = brd("thin", "thin", "thin", "thin")
B_NONE = brd()
B_MED_L = brd("medium", None, None, None)


def sw(ws, r, c, val=None, f=None, a=None, b=None):
    cell = ws.cell(row=r, column=c, value=val)
    if f: cell.font = f
    if a: cell.alignment = a
    if b: cell.border = b
    return cell


def mw(ws, r1, c1, r2, c2, val=None, f=None, a=None, b=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return sw(ws, r1, c1, val, f, a, b)


# =============================================================================
# Register-number helpers
# =============================================================================

_BM = {
    "CS": "CSE", "EC": "ECE", "CE": "CE", "IT": "IT",
    "ME": "ME", "EE": "EEE", "AE": "AE", "ER": "ERE", "ECE": "ERE"
}


def branch_of(reg):
    if not reg: return None
    m = re.search(r"[A-Z]{3}\d{2}([A-Z]+)\d+$", reg.upper())
    return _BM.get(m.group(1)) if m else None


def is_supply(reg):
    return bool(reg and reg.upper().startswith("LL"))


def make_range_str(regs):
    if not regs: return ""
    cl = sorted(set(regs))
    if len(cl) == 1: return cl[0] + " (1)"
    groups, g = [], [cl[0]]
    for reg in cl[1:]:
        pm = re.search(r"\d+$", g[-1])
        cm = re.search(r"\d+$", reg)
        if (pm and cm and g[-1][:pm.start()] == reg[:cm.start()]
                and int(cm.group()) == int(pm.group()) + 1):
            g.append(reg)
        else:
            groups.append(g)
            g = [reg]
    groups.append(g)
    parts = []
    for grp in groups:
        nm = re.search(r"\d+$", grp[-1])
        if len(grp) > 1 and nm:
            # Get the prefix (everything before the number)
            prefix = grp[0][:nm.start()]
            # Get the first number
            first_num = re.search(r"\d+$", grp[0]).group()
            if prefix == grp[0][:nm.start()]:
                # If all have same prefix, show range
                parts.append(f"{grp[0]}-{nm.group()}")
            else:
                # Otherwise list individually
                parts.append(", ".join(grp))
        else:
            parts.append(grp[0])
    return ", ".join(parts) + f" ({len(cl)})"


# =============================================================================
# Parse TXT and extract metadata
# =============================================================================

def extract_metadata_from_txt(text, folder_path=None, exam_date_param=None):
    """Extract metadata like exam title, date, institution from TXT content"""
    metadata = {
        'exam_title': 'SEATING ARRANGEMENT',
        'exam_date': '',
        'institution': 'LBS INSTITUTE OF TECHNOLOGY FOR WOMEN, Poojappura, Thiruvananthapuram',
        'semester': 'S4',
        'scheme': '2019',
        'session': 'FN',
        'absentee_header': 'ABSENTEE STATEMENT',
        'door_exam': '',
        'day': '',
        'month': '',
        'year': '',
        'date_display': ''
    }

    # FIRST PRIORITY: Use exam_date_param if provided
    if exam_date_param:
        # exam_date_param format: YYYY-MM-DD
        date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', exam_date_param)
        if date_match:
            year = date_match.group(1)
            month = date_match.group(2)
            day = date_match.group(3)
            metadata['exam_date'] = f"{day}.{month}.{year}"
            metadata['day'] = day
            metadata['month'] = month
            metadata['year'] = year
            metadata['date_display'] = f"{day}.{month}.{year} {metadata['session']}"

    # SECOND PRIORITY: Try to extract from folder path
    if folder_path and not metadata['date_display']:
        date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', folder_path)
        if date_match:
            year = date_match.group(1)
            month = date_match.group(2)
            day = date_match.group(3)
            metadata['exam_date'] = f"{day}.{month}.{year}"
            metadata['day'] = day
            metadata['month'] = month
            metadata['year'] = year
            metadata['date_display'] = f"{day}.{month}.{year} {metadata['session']}"

    # Try to extract exam title with date pattern from text
    title_match = re.search(r'(S[1-8]\([^)]+\)\s+\d{4}\s+SCHEME\s+[A-Z]+\s+EXAM\s+SEATING\s+ARRANGEMENT)', text, re.IGNORECASE)
    if title_match:
        metadata['exam_title'] = title_match.group(1).strip()

    # Try to extract semester and scheme
    sem_match = re.search(r'(S[1-8])\(([^)]+)\)\s+(\d{4})\s+SCHEME', text, re.IGNORECASE)
    if sem_match:
        metadata['semester'] = sem_match.group(1)
        metadata['scheme'] = sem_match.group(3)
        if not metadata['exam_title'] or metadata['exam_title'] == 'SEATING ARRANGEMENT':
            metadata['exam_title'] = f"{metadata['semester']}({sem_match.group(2)}) {metadata['scheme']} SCHEME B.TECH DEGREE EXAM SEATING ARRANGEMENT"

    # Try to extract session (FN/AN)
    session_match = re.search(r'(FN|AN)\b', text)
    if session_match:
        metadata['session'] = session_match.group(1)
        # Update date_display with correct session if we have a date
        if metadata['exam_date']:
            metadata['date_display'] = f"{metadata['exam_date']} {metadata['session']}"

    # Try to extract door exam line
    door_match = re.search(r'(S[1-8]\s*\([^)]+\)\s*\d{4}\s+B\.TECH\s+DEGREE\s+EXAMINATION)', text, re.IGNORECASE)
    if door_match:
        metadata['door_exam'] = door_match.group(1).strip()

    return metadata

def parse_txt(txt_content, metadata):
    """Parse TXT content and extract room data"""
    subjects = {}
    for m in re.finditer(
            r"^([A-Z][A-Z /\-]+\(\s*([A-Z0-9]+)\s*\))\s*:\s*(\d+)\s*students",
            txt_content, re.MULTILINE):
        subjects[m.group(2).strip()] = (m.group(1).strip(), int(m.group(3)))

    raw = re.split(r"={10,}", txt_content)
    hidx = [i for i, b in enumerate(raw) if re.search(r"Room\d+\s*\|\s*TOTAL:", b)]
    rooms = []

    for idx in hidx:
        block = raw[idx] + ("\n" + raw[idx + 1] if idx + 1 < len(raw) else "")
        hm = re.search(r"Room(\d+)\s*\|\s*TOTAL:\s*(\d+)", block)
        if not hm: continue

        rows_data = []
        for rm in re.finditer(r"^\s*(\d+)\s+\|(.*)", block, re.MULTILINE):
            vals = []
            for cv in rm.group(2).split("|"):
                v = cv.strip()
                vals.append(None if v in ("--", "") else v)
            while vals and vals[-1] is None:
                vals.pop()
            rows_data.append((int(rm.group(1)), vals))

        rsubj = {}
        sl = re.search(r"Subjects:\s*(.*)", block)
        if sl:
            for sm in re.finditer(
                    r"(S\d+\w+):\s*[\w ]+\(\s*([A-Z0-9]+)\s*\)", sl.group(1)):
                rsubj[sm.group(1)] = sm.group(2)

        rooms.append({
            "room_num": int(hm.group(1)),
            "room_name": f"Room{int(hm.group(1))}",
            "total": int(hm.group(2)),
            "rows": rows_data,
            "subjects": rsubj,
        })

    return {"subjects": subjects}, rooms


def room_students(room):
    bd = defaultdict(lambda: {"r": [], "s": []})
    for _, vals in room["rows"]:
        for v in vals:
            if v:
                br = branch_of(v)
                if br:
                    bd[br]["s" if is_supply(v) else "r"].append(v)
    for br in bd:
        bd[br]["r"] = sorted(set(bd[br]["r"]))
        bd[br]["s"] = sorted(set(bd[br]["s"]))
    return dict(bd)


def ordered_br(bd, order=None):
    DEPT_ORDER = ["CE", "CSE", "ECE", "IT", "ERE", "ME", "EEE", "AE", "AEI"]
    o = order or DEPT_ORDER
    return [b for b in o if b in bd] + [b for b in sorted(bd) if b not in o]


_T2B = {
    "EC": "ECE", "CS": "CSE", "CE": "CE", "IT": "IT",
    "ER": "ERE", "ME": "ME", "EE": "EEE", "AE": "AE", "AEI": "AEI"
}


def subj_map(room, meta):
    res = {}
    for tag, code in room["subjects"].items():
        full, _ = meta["subjects"].get(code, (code, 0))
        bt = re.sub(r"^S\d+", "", tag)
        res[_T2B.get(bt, bt)] = full
    return res


# =============================================================================
# Build sheets
# =============================================================================

# Room to hall mapping (can be customized)
def get_room_to_hall_map():
    return {
        "Room1": "108", "Room2": "109", "Room3": "111", "Room4": "112",
        "Room5": "208", "Room6": "209", "Room7": "211", "Room8": "212",
        "Room9": "308", "Room10": "309", "Room11": "311", "Room12": "312",
    }


# MASTER sheet border patterns
_MDB = {
    "CE": ("double", "medium", "double", "medium"),
    "CSE": (None, "medium", None, None),
    "ECE": ("medium", "thin", "medium", "thin"),
    "IT": ("thin", "medium", "thin", "medium"),
    "ERE": ("medium", "medium", None, None),
    "AEI": ("medium", "thin", "medium", "thin"),
    "_": ("thin", "thin", "thin", "thin"),
}


def build_master(ws, meta, rooms, hall_map):
    """Build MASTER sheet with proper formatting"""
    # Set column widths
    ws.column_dimensions["A"].width = 30.14
    ws.column_dimensions["B"].width = 198.86
    ws.column_dimensions["C"].width = 33.0

    # Fonts
    F36 = fnt("Cambria", 36, True)
    F26 = fnt("Cambria", 26, True)
    F24 = fnt("Cambria", 24, True)
    F28 = fnt("Cambria", 28, True)
    AC = aln("center", "center")
    AL = aln("left", "center")

    # Row heights
    ws.row_dimensions[1].height = 73.5
    ws.row_dimensions[2].height = 63.0
    ws.row_dimensions[3].height = 61.5
    ws.row_dimensions[4].height = 64.5

    # Rows 1-3: institution / university / exam title
    mw(ws, 1, 1, 1, 3, meta.get('institution', ''),
       F36, AC, brd("medium", "thin", "medium", "thin"))
    mw(ws, 2, 1, 2, 3, "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY (APJAKTU)",
       F26, AC, brd("medium", "medium", None, None))

    # Row 3: exam title with date
    exam_title = meta.get('exam_title', 'SEATING ARRANGEMENT')
    date_display = meta.get('date_display', '')
    if date_display:
        mw(ws, 3, 1, 3, 3, f"{exam_title} {date_display}",
           F24, AC, brd("medium", "medium", "thin", "double"))
    else:
        mw(ws, 3, 1, 3, 3, exam_title,
           F24, AC, brd("medium", "medium", "thin", "double"))

    # Row 4: column headers
    sw(ws, 4, 1, "BRANCH", F26, AC, brd("medium", "thin", "double", "double"))
    sw(ws, 4, 2, "REGISTER NUMBERS", F26, AC, brd("thin", "thin", "double", "double"))
    sw(ws, 4, 3, "HALL NO", F26, AC, brd("thin", "medium", "double", "double"))

    # Collect: branch -> hall -> students
    bh = defaultdict(lambda: defaultdict(lambda: {"r": [], "s": []}))
    for room in rooms:
        hall = hall_map.get(room["room_name"], room["room_name"])
        for br, data in room_students(room).items():
            bh[br][hall]["r"].extend(data["r"])
            bh[br][hall]["s"].extend(data["s"])
    for br in bh:
        for h in bh[br]:
            bh[br][h]["r"] = sorted(set(bh[br][h]["r"]))
            bh[br][h]["s"] = sorted(set(bh[br][h]["s"]))

    row = 5
    for br in ordered_br(bh):
        halls = sorted(bh[br].keys(), key=lambda x: int(x) if x.isdigit() else 999)
        br_rows = [(h, make_range_str(bh[br][h]["r"] + bh[br][h]["s"]))
                   for h in halls if bh[br][h]["r"] or bh[br][h]["s"]]
        if not br_rows:
            continue
        span = len(br_rows)
        for i in range(span):
            ws.row_dimensions[row + i].height = 80.1

        at, ab, bft, blb = _MDB.get(br, _MDB["_"])
        if row == 5:
            at = "double"
            bft = "double"
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row + span - 1, end_column=1)

        # Format branch cell
        branch_cell = sw(ws, row, 1, br, F28, AC, brd("medium", "thin", at, ab))
        if span > 1:
            branch_cell.alignment = aln("center", "center")

        for i, (hall, rng) in enumerate(br_rows):
            r = row + i
            top = bft if i == 0 else ("thin" if bft else None)
            bot = blb if i == span - 1 else "thin"
            # Register numbers column - left aligned
            sw(ws, r, 2, rng, F28, AL, brd("thin", "thin", top, bot))
            # Hall number column - center aligned
            sw(ws, r, 3, int(hall) if hall.isdigit() else hall,
               F28, AC, brd("thin", "medium", top, bot))
        row += span


# Hall sheet layout - Updated for better formatting
_HL = [
    ("AEI", 8, 11, 2, "thin", "medium", []),
    ("CE", 8, 11, 2, "thin", "medium", []),
    ("CSE", 15, 19, 3, None, "medium", [(22, 22)]),
    ("ECE", 25, 28, 2, None, "medium", [(29, 29)]),
    ("ERE", 32, 35, 2, "thin", "medium", []),
    ("IT", 39, 42, 2, "thin", "medium", []),
]


def build_hall_sheet(wb, room, meta, hall_map, room_index):
    """Build individual hall sheet with proper date formatting"""
    hall_no = hall_map.get(room["room_name"], room["room_name"])
    ws = wb.create_sheet(f"{hall_no}")

    # Column widths - matching the original Excel
    ws.column_dimensions['A'].width = 8.71
    ws.column_dimensions['B'].width = 7.57
    ws.column_dimensions['C'].width = 9.43
    ws.column_dimensions['D'].width = 8.71
    ws.column_dimensions['E'].width = 8.71
    ws.column_dimensions['F'].width = 15.57
    ws.column_dimensions['G'].width = 7.71
    ws.column_dimensions['H'].width = 8.71
    ws.column_dimensions['I'].width = 3.14

    # Row heights
    for r in range(1, 51):
        ws.row_dimensions[r].height = 15
    ws.row_dimensions[21].height = 7.5
    ws.row_dimensions[29].height = 9.75
    ws.row_dimensions[43].height = 10.5
    ws.row_dimensions[46].height = 24.95
    ws.row_dimensions[47].height = 20.1
    ws.row_dimensions[48].height = 20.1
    ws.row_dimensions[50].height = 15.75

    F12b = fnt("Cambria", 12, True)
    F12 = fnt("Cambria", 12, False)
    F11b = fnt("Cambria", 11, True)
    F11 = fnt("Cambria", 11, False)
    F9 = fnt("Cambria", 9, False)
    AC = aln("center", "center")
    AL = aln("left", "center")
    AR = aln("right", "center")

    # Row 1: Institution
    mw(ws, 1, 1, 1, 9,
       meta.get('institution', 'LBS INSTITUTE OF TECHNOLOGY FOR WOMEN, Poojappura, Thiruvananthapuram'),
       F12b, AC, brd("medium", "medium", "medium", None))

    # Row 2: Absentee header with month
    semester = meta.get('semester', 'S4')
    scheme = meta.get('scheme', '2019')

    # Extract month from date if available
    month_name = "APRIL"
    if meta.get('month'):
        month_num = int(meta.get('month'))
        month_names = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                       'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
        if 1 <= month_num <= 12:
            month_name = month_names[month_num - 1]

    year_short = meta.get('year', '25')[-2:] if meta.get('year') else '25'
    absentee_hdr = f"ABSENTEE STATEMENT, KTU-{semester}(R,S) {scheme} SCHEME, {month_name} '{year_short}"
    mw(ws, 2, 1, 2, 9, absentee_hdr,
       F11b, AC, brd("medium", "medium", None, None))

    # Row 3: blank
    mw(ws, 3, 1, 3, 9, None, None, None, None)

    # Rows 4-5: Date and Hall No
    date_display = meta.get('date_display', '')
    if not date_display:
        exam_date = meta.get('exam_date', '')
        if exam_date:
            session_val = meta.get('session', 'FN')
            date_display = f"{exam_date} {session_val}"

    if not date_display:
        date_display = "Date not found"

    # Merge cells for date row
    mw(ws, 4, 1, 5, 1, "DATE : ", F12, AR, brd("medium", "thin", "thin", "thin"))
    mw(ws, 4, 2, 5, 3, date_display, F12, AL, brd("thin", "thin", "thin", "thin"))
    mw(ws, 4, 7, 5, 8, "Hall No.", F12, AC, brd("thin", "thin", "thin", "thin"))
    mw(ws, 4, 9, 5, 9, hall_no, F12, AC, brd("thin", "medium", "thin", "thin"))

    # Clear cells in between
    for rr in [4, 5]:
        for cc in [4, 5, 6]:
            sw(ws, rr, cc, None, F11, AC, B_NONE)

    # Row 7: left blank merge + right label
    mw(ws, 7, 1, 7, 6, None, F11, AC, brd("medium", "thin", None, "thin"))
    mw(ws, 7, 7, 7, 9, "REGISTER NUMBER OF ABSENTEES",
       F9, AC, brd("thin", "thin", None, "thin"))

    # Per-dept sections
    bd = room_students(room)
    sm = subj_map(room, meta)

    for (dept, hdr_row, data_row, data_nrows,
         r_brd_L, r_brd_R, extra_rights) in _HL:

        if dept not in bd and dept not in ["CE", "CSE", "ECE", "IT", "ERE", "AEI"]:
            continue

        # Header row
        sw(ws, hdr_row, 1, dept,
           F12b, aln("center", None), brd("medium", "thin", "thin", "thin"))

        subject_text = sm.get(dept, "")
        mw(ws, hdr_row, 2, hdr_row, 6,
           f"Subject: {subject_text}" if subject_text else "Subject:",
           F12, AL, brd("thin", "thin", "thin", "thin"))
        mw(ws, hdr_row, 7, hdr_row, 9,
           dept, F12b, aln("center", None), brd("thin", "thin", "thin", "thin"))

        # Data area left - merged cells
        mw(ws, data_row, 1, data_row + data_nrows - 1, 6,
           None, F12, aln("center", "center", True),
           brd("medium", "thin", None, None))

        # Right-side cells
        r_fnt = F12 if dept == "IT" else F12b
        for rr in range(data_row, data_row + data_nrows):
            sw(ws, rr, 7, None, r_fnt, None, brd(r_brd_L, None, None, None))
            sw(ws, rr, 8, None, r_fnt, None, brd(None, None, None, None))
            sw(ws, rr, 9, None, r_fnt, None, brd(None, r_brd_R, None, None))

        for (er1, er2) in extra_rights:
            mw(ws, er1, 7, er2, 9, None, F12b, None, brd(None, "medium", None, None))

        # Fill data value
        data = bd.get(dept) if dept in bd else None
        if data and (data["r"] or data["s"]):
            parts = []
            if data["r"]:
                parts.append(make_range_str(data["r"]))
            if data["s"]:
                parts.append(make_range_str(data["s"]))
            val = "\n".join(parts)
            wrap = len(parts) > 1
        else:
            val = "NIL"
            wrap = False

        c = ws.cell(data_row, 1)
        c.value = val
        c.font = F12
        c.alignment = aln("center", "center", wrap)
        c.border = brd("medium", "thin", None, None)

    # Footer rows
    mw(ws, 46, 1, 46, 3, "No. of Question Papers received  ",
       F11, AC, brd("medium", "thin", "thin", "thin"))
    sw(ws, 46, 4, None, F11, AC, brd("thin", None, "thin", "thin"))
    mw(ws, 46, 5, 46, 6, "Issued ", F11, AC, brd("thin", "thin", "thin", "thin"))
    sw(ws, 46, 7, None, F11, AC, brd("thin", "thin", "thin", "thin"))
    mw(ws, 46, 8, 46, 9, "Balance", F11, AC, brd("thin", "thin", "thin", "thin"))
    sw(ws, 46, 9, None, F11, AC, brd("thin", "medium", "thin", "thin"))

    mw(ws, 47, 1, 47, 3, "No. of Answer Books received ",
       F11, AC, brd("medium", "thin", "thin", "thin"))
    sw(ws, 47, 4, None, F11, AC, brd("thin", None, "thin", "thin"))
    mw(ws, 47, 5, 47, 6, "Issued ", F11, AC, brd("thin", "thin", "thin", "thin"))
    sw(ws, 47, 7, None, F11, AC, brd("thin", "thin", "thin", "thin"))
    mw(ws, 47, 8, 47, 9, "Balance", F11, AC, brd("thin", "thin", "thin", "thin"))
    sw(ws, 47, 9, None, F11, AC, brd("thin", "medium", "thin", "thin"))

    mw(ws, 48, 1, 48, 4, "Invigilator's Name & Designation",
       F11b, AC, brd("medium", "thin", "thin", "thin"))
    mw(ws, 48, 5, 48, 7, "Department", F11b, AC, brd("thin", "thin", "thin", "thin"))
    mw(ws, 48, 8, 48, 9, "Dated Signature", F11b, AC, brd("thin", "thin", "thin", "thin"))

    mw(ws, 49, 1, 50, 4, None, F11, AC, brd("medium", "thin", "thin", "thin"))
    mw(ws, 49, 5, 50, 7, None, F11, AC, brd("thin", "thin", "thin", "thin"))
    mw(ws, 49, 8, 50, 9, None, F11, AC, brd("thin", "thin", "thin", "thin"))


def build_door_sheet(wb, rooms, hall_map, metadata):
    """Build DOOR sheet with SEATING PLAN"""
    ws = wb.create_sheet("SEATING PLAN")

    col_widths = {
        'A': 1.14, 'B': 17.29, 'C': 18.0, 'D': 2.43,
        'E': 18.0, 'F': 2.14, 'G': 19.14, 'H': 17.0,
        'I': 1.57, 'J': 2.14
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    F14b = fnt("Calibri", 14, True)
    F18 = fnt("Calibri", 18, False)
    F11b = fnt("Cambria", 11, True)
    F11 = fnt("Cambria", 11, False)
    AC = aln("center", "center")
    AL = aln("left", "center")
    ALT = aln("left", "top", True)

    _DOOR_COLS = [2, 3, 5, 7, 8]

    def _door_col_a(ws, r, extra_top=False, extra_bot=False):
        f = fnt("Calibri", 11, False)
        t = "medium" if extra_top else None
        b = "medium" if extra_bot else None
        sw(ws, r, 1, None, f, None, brd("medium", None, t, b))

    cur = 1

    for room in rooms:
        hall_no = hall_map.get(room["room_name"], room["room_name"])
        rows = room["rows"]
        block_start = cur

        # Header
        _door_col_a(ws, cur, extra_top=(cur == block_start))
        mw(ws, cur, 2, cur, 8,
           "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY",
           F14b, AC, brd(None, None, "medium", None))
        ws.row_dimensions[cur].height = 18.75
        cur += 1

        _door_col_a(ws, cur)
        door_exam = metadata.get(
            'door_exam') or f"{metadata.get('semester', 'S4')} (R,S){metadata.get('scheme', '2019')} B.TECH DEGREE EXAMINATION"
        mw(ws, cur, 2, cur, 8, door_exam, F14b, AC, B_NONE)
        ws.row_dimensions[cur].height = 18.75
        cur += 1

        _door_col_a(ws, cur)
        mw(ws, cur, 3, cur, 7, "SEATING PLAN", F14b, AC, B_NONE)
        sw(ws, cur, 8, None, F14b, None, B_NONE)
        ws.row_dimensions[cur].height = 18.75
        cur += 1

        _door_col_a(ws, cur)
        date_display = metadata.get('date_display', '')
        mw(ws, cur, 2, cur, 3, date_display, F18, AL, B_ALL)
        mw(ws, cur, 7, cur, 8, f"HALL : {hall_no}", F18, AL, B_ALL)
        ws.row_dimensions[cur].height = 23.25
        cur += 1

        _door_col_a(ws, cur)
        cur += 1

        # Column numbers
        _door_col_a(ws, cur)
        for idx, cp in enumerate(_DOOR_COLS, 1):
            sw(ws, cur, cp, str(idx), F14b, AC, B_ALL)
        sw(ws, cur, 4, None, fnt("Calibri", 11), None, B_NONE)
        sw(ws, cur, 6, None, fnt("Calibri", 11), None, B_NONE)
        ws.row_dimensions[cur].height = 18.75
        cur += 1

        # Seating grid
        max_rn = max((rn for rn, _ in rows), default=0)
        col_data = defaultdict(dict)
        for rn, vals in rows:
            for ci, v in enumerate(vals):
                if v:
                    col_data[ci][rn] = v

        for rnum in range(1, max_rn + 1):
            if not any(col_data[ci].get(rnum) for ci in range(5)):
                continue
            _door_col_a(ws, cur)
            for ci, cp in enumerate(_DOOR_COLS):
                v = col_data[ci].get(rnum)
                if v:
                    sw(ws, cur, cp, v, F14b, AC, B_ALL)
            sw(ws, cur, 4, None, F14b, None, B_NONE)
            sw(ws, cur, 6, None, F14b, None, B_NONE)
            ws.row_dimensions[cur].height = 18.75
            cur += 1

        # Blank row
        _door_col_a(ws, cur)
        cur += 1

        # Branch summary
        bd = room_students(room)
        reg_bd = {br: data["r"] for br, data in bd.items() if data["r"]}
        sup_bd = {br: data["s"] for br, data in bd.items() if data["s"]}

        for br in ordered_br(bd):
            if br not in reg_bd:
                continue
            _door_col_a(ws, cur)
            sw(ws, cur, 2, br, F11b, AL, B_ALL)
            mw(ws, cur, 3, cur, 8, make_range_str(reg_bd[br]), F11, AL, B_ALL)
            ws.row_dimensions[cur].height = 20.1
            cur += 1

        if sup_bd:
            parts = [f"{br} - {make_range_str(sup_bd[br])}"
                     for br in ordered_br(bd) if br in sup_bd]
            _door_col_a(ws, cur)
            _door_col_a(ws, cur + 1)
            mw(ws, cur, 2, cur + 1, 2, "Supply", F11b, AL, B_ALL)
            mw(ws, cur, 3, cur + 1, 8, ", ".join(parts), F11, ALT, B_ALL)
            ws.row_dimensions[cur].height = 14.25
            ws.row_dimensions[cur + 1].height = 32.25
            cur += 2

        # Gap rows
        gap_heights = [17.25, 17.25, 17.25, 17.25, 15.75]
        for gi, gh in enumerate(gap_heights):
            is_last_gap = (gi == len(gap_heights) - 1)
            _door_col_a(ws, cur, extra_bot=is_last_gap)
            if is_last_gap:
                for c in range(2, 9):
                    sw(ws, cur, c, None, fnt("Calibri", 11), None,
                       brd(None, None, None, "medium"))
            ws.row_dimensions[cur].height = gh
            cur += 1


# =============================================================================
# Main conversion function
# =============================================================================

def convert_txt_to_excel(txt_file_path, output_excel_path=None, folder_path=None, exam_date=None):
    """
    Main function to convert TXT file to Excel
    Returns path to generated Excel file and metadata
    """
    try:
        # Read TXT file
        with open(txt_file_path, 'r', encoding='utf-8', errors='replace') as f:
            txt_content = f.read()

        # Extract metadata with folder path and exam_date for date extraction
        metadata = extract_metadata_from_txt(txt_content, folder_path, exam_date)

        # Parse TXT content
        meta, rooms = parse_txt(txt_content, metadata)

        # IMPORTANT: Merge metadata into meta so it's available in build_hall_sheet
        meta.update(metadata)

        if not rooms:
            return {'success': False, 'error': 'No rooms found in TXT file'}

        # Generate output path if not provided
        if output_excel_path is None:
            output_excel_path = txt_file_path.replace('.txt', '_seating_output.xlsx')

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Get room to hall mapping
        room_to_hall = get_room_to_hall_map()

        # Build sheets
        build_master(wb.create_sheet("MASTER"), meta, rooms, room_to_hall)

        for i, room in enumerate(rooms):
            build_hall_sheet(wb, room, meta, room_to_hall, i)

        build_door_sheet(wb, rooms, room_to_hall, metadata)

        # Save workbook
        try:
            wb.save(output_excel_path)
        except PermissionError:
            base, ext = os.path.splitext(output_excel_path)
            for i in range(1, 21):
                alt_path = f"{base}_{i}{ext}"
                try:
                    wb.save(alt_path)
                    output_excel_path = alt_path
                    break
                except PermissionError:
                    continue
            else:
                return {'success': False, 'error': 'Cannot save file - permission denied'}

        return {
            'success': True,
            'excel_path': output_excel_path,
            'metadata': metadata,
            'rooms_count': len(rooms)
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}

# For standalone usage
if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        txt_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        folder_path = sys.argv[3] if len(sys.argv) > 3 else None
        exam_date = sys.argv[4] if len(sys.argv) > 4 else None
        result = convert_txt_to_excel(txt_file, output_file, folder_path, exam_date)
        if result['success']:
            print(f"Success! Excel file saved to: {result['excel_path']}")
            print(f"Metadata extracted: {result['metadata']}")
        else:
            print(f"Error: {result['error']}")
    else:
        print("Usage: python txt_to_excel_converter.py <input_txt_file> [output_excel_file] [folder_path] [exam_date]")
