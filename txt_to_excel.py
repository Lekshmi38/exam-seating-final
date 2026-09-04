"""
Seating Arrangement TXT -> Excel Converter  (pixel-perfect match to original)
=============================================================================
Edit the CONFIGURATION block below, then run:
    python txt_to_excel.py
"""

import re, sys, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# =============================================================================
# CONFIGURATION  <-- Edit these, then just run the script
# =============================================================================

TXT_FILE     = r"C:\Users\User\Downloads\seating_report1.txt"
OUTPUT_EXCEL = r"C:\Users\User\Downloads\seating_output.xlsx"

INSTITUTION  = "LBS INSTITUTE OF TECHNOLOGY FOR WOMEN, Poojappura, Thiruvananthapuram"
EXAM_TITLE   = "S7 2019 SCHEME B.TECH DEGREE EXAM SEATING ARRANGEMENT 25.04.2025 FN"
ABSENTEE_HDR = "ABSENTEE STATEMENT, KTU-S7 2019 SCHEME, APRIL '25"
EXAM_DATE    = "25.04.2025 FN"
DOOR_EXAM    = "S7 (R, S)2019 B.TECH DEGREE EXAMINATION"

ROOM_TO_HALL = {
    "Room1":  "108", "Room2":  "109", "Room3":  "111", "Room4":  "112",
    "Room5":  "208", "Room6":  "209", "Room7":  "211", "Room8":  "212",
    "Room9":  "308", "Room10": "309", "Room11": "311", "Room12": "312",
}

# Order depts appear in MASTER / DOOR summaries
DEPT_ORDER = ["CE", "CSE", "ECE", "IT", "ERE", "ME", "EEE", "AE"]
# Fixed 5 dept sections on every hall absentee sheet (matches original layout)
HALL_DEPTS  = ["CE", "CSE", "ECE", "ERE", "IT"]

# =============================================================================
# Low-level style helpers
# =============================================================================

def S(v):  return Side(style=v) if v else Side(style=None)
def brd(L=None, R=None, T=None, B=None):
    return Border(left=S(L), right=S(R), top=S(T), bottom=S(B))
def fnt(name="Cambria", size=12, bold=False):
    return Font(name=name, size=size, bold=bold)
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Re-usable border constants
B_ALL   = brd("thin","thin","thin","thin")
B_NONE  = brd()
B_MED_L = brd("medium",None,None,None)     # left-medium only (DOOR col-A style)

def sw(ws, r, c, val=None, f=None, a=None, b=None):
    cell = ws.cell(row=r, column=c, value=val)
    if f: cell.font      = f
    if a: cell.alignment = a
    if b: cell.border    = b
    return cell

def mw(ws, r1, c1, r2, c2, val=None, f=None, a=None, b=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return sw(ws, r1, c1, val, f, a, b)

# =============================================================================
# Register-number helpers
# =============================================================================

_BM = {"CS":"CSE","EC":"ECE","CE":"CE","IT":"IT",
       "ME":"ME","EE":"EEE","AE":"AE","ER":"ERE","ECE":"ERE"}

def branch_of(reg):
    if not reg: return None
    m = re.search(r"[A-Z]{3}\d{2}([A-Z]+)\d+$", reg.upper())
    return _BM.get(m.group(1)) if m else None

def is_supply(reg):
    return bool(reg and reg.upper().startswith("LL"))

def make_range_str(regs):
    if not regs: return ""
    cl = sorted(set(regs))
    if len(cl) == 1: return cl[0] + " (1)"
    groups, g = [], [cl[0]]
    for reg in cl[1:]:
        pm = re.search(r"\d+$", g[-1]); cm = re.search(r"\d+$", reg)
        if (pm and cm and g[-1][:pm.start()]==reg[:cm.start()]
                and int(cm.group())==int(pm.group())+1):
            g.append(reg)
        else:
            groups.append(g); g = [reg]
    groups.append(g)
    parts = []
    for grp in groups:
        nm = re.search(r"\d+$", grp[-1])
        parts.append(grp[0]+"-"+nm.group() if len(grp)>1 and nm else grp[0])
    return ", ".join(parts) + f" ({len(cl)})"

# =============================================================================
# Parse TXT
# =============================================================================

def parse_txt(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    subjects = {}
    for m in re.finditer(
            r"^([A-Z][A-Z /\-]+\(\s*([A-Z0-9]+)\s*\))\s*:\s*(\d+)\s*students",
            text, re.MULTILINE):
        subjects[m.group(2).strip()] = (m.group(1).strip(), int(m.group(3)))
    raw  = re.split(r"={10,}", text)
    hidx = [i for i,b in enumerate(raw) if re.search(r"Room\d+\s*\|\s*TOTAL:",b)]
    rooms = []
    for idx in hidx:
        block = raw[idx] + ("\n"+raw[idx+1] if idx+1<len(raw) else "")
        hm = re.search(r"Room(\d+)\s*\|\s*TOTAL:\s*(\d+)", block)
        if not hm: continue
        rows_data = []
        for rm in re.finditer(r"^\s*(\d+)\s+\|(.*)", block, re.MULTILINE):
            vals = []
            for cv in rm.group(2).split("|"):
                v=cv.strip(); vals.append(None if v in ("--","") else v)
            while vals and vals[-1] is None: vals.pop()
            rows_data.append((int(rm.group(1)), vals))
        rsubj = {}
        sl = re.search(r"Subjects:\s*(.*)", block)
        if sl:
            for sm in re.finditer(
                    r"(S\d+\w+):\s*[\w ]+\(\s*([A-Z0-9]+)\s*\)", sl.group(1)):
                rsubj[sm.group(1)] = sm.group(2)
        rooms.append({
            "room_num":  int(hm.group(1)),
            "room_name": f"Room{int(hm.group(1))}",
            "total":     int(hm.group(2)),
            "rows":      rows_data,
            "subjects":  rsubj,
        })
    return {"subjects": subjects}, rooms

def room_students(room):
    bd = defaultdict(lambda: {"r":[], "s":[]})
    for _,vals in room["rows"]:
        for v in vals:
            if v:
                br = branch_of(v)
                if br: bd[br]["s" if is_supply(v) else "r"].append(v)
    for br in bd:
        bd[br]["r"] = sorted(set(bd[br]["r"]))
        bd[br]["s"] = sorted(set(bd[br]["s"]))
    return dict(bd)

def ordered_br(bd, order=None):
    o = order or DEPT_ORDER
    return [b for b in o if b in bd]+[b for b in sorted(bd) if b not in o]

_T2B = {"EC":"ECE","CS":"CSE","CE":"CE","IT":"IT",
        "ER":"ERE","ME":"ME","EE":"EEE","AE":"AE"}

def subj_map(room, meta):
    res = {}
    for tag,code in room["subjects"].items():
        full,_ = meta["subjects"].get(code,(code,0))
        bt = re.sub(r"^S\d+","",tag)
        res[_T2B.get(bt,bt)] = full
    return res

# =============================================================================
# MASTER sheet
# Exact border pattern from original (verified cell-by-cell):
#   dept  | A_top    A_bot   | B/C_first_top  B/C_last_bot
#   CE    | double   medium  | double         medium
#   CSE   | None     medium  | None           None
#   ECE   | medium   thin    | medium         thin
#   IT    | thin     medium  | thin           medium
#   ERE   | medium   medium  | None           None
#   other | thin     thin    | thin           thin
# =============================================================================

_MDB = {                       # (A_top, A_bot, B_first_top, B_last_bot)
    "CE":  ("double","medium","double","medium"),
    "CSE": (None,    "medium", None,    None   ),
    "ECE": ("medium","thin",  "medium","thin"  ),
    "IT":  ("thin",  "medium","thin",  "medium"),
    "ERE": ("medium","medium", None,    None   ),
    "_":   ("thin",  "thin",  "thin",  "thin"  ),
}

def build_master(wb, meta, rooms, hall_map):
    ws = wb.create_sheet("MASTER")
    ws.column_dimensions["A"].width = 30.14
    ws.column_dimensions["B"].width = 198.86
    ws.column_dimensions["C"].width = 33.0
    ws.column_dimensions["D"].width = 9.0

    F36=fnt("Cambria",36,True); F26=fnt("Cambria",26,True)
    F24=fnt("Cambria",24,True); F28=fnt("Cambria",28,True)
    AC=aln("center","center");  AL=aln("left","center")

    ws.row_dimensions[1].height = 73.5
    ws.row_dimensions[2].height = 63.0
    ws.row_dimensions[3].height = 61.5
    ws.row_dimensions[4].height = 64.5

    # Rows 1-3: institution / university / exam title
    mw(ws,1,1,1,3, INSTITUTION,
       F36, AC, brd("medium","thin","medium","thin"))
    mw(ws,2,1,2,3, "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY (APJAKTU)",
       F26, AC, brd("medium","medium",None,None))
    mw(ws,3,1,3,3, EXAM_TITLE,
       F24, AC, brd("medium","medium","thin","double"))

    # Row 4: column headers
    sw(ws,4,1,"BRANCH",           F26,AC,brd("medium","thin","double","double"))
    sw(ws,4,2,"REGISTER NUMBERS", F26,AC,brd("thin","thin","double","double"))
    sw(ws,4,3,"HALL NO",          F26,AC,brd("thin","medium","double","double"))

    # Collect: branch -> hall -> students
    bh = defaultdict(lambda: defaultdict(lambda:{"r":[],"s":[]}))
    for room in rooms:
        hall = hall_map.get(room["room_name"], room["room_name"])
        for br,data in room_students(room).items():
            bh[br][hall]["r"].extend(data["r"])
            bh[br][hall]["s"].extend(data["s"])
    for br in bh:
        for h in bh[br]:
            bh[br][h]["r"] = sorted(set(bh[br][h]["r"]))
            bh[br][h]["s"] = sorted(set(bh[br][h]["s"]))

    row = 5
    for br in ordered_br(bh):
        halls = sorted(bh[br].keys(), key=lambda x: int(x))
        br_rows = [(h, make_range_str(bh[br][h]["r"]+bh[br][h]["s"]))
                   for h in halls if bh[br][h]["r"] or bh[br][h]["s"]]
        if not br_rows: continue
        span = len(br_rows)
        for i in range(span): ws.row_dimensions[row+i].height = 80.1

        at, ab, bft, blb = _MDB.get(br, _MDB["_"])
        # First dept group rendered always gets double-top (original rule)
        if row == 5:
            at  = "double"
            bft = "double"
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row+span-1, end_column=1)
        sw(ws, row, 1, br, F28, AC, brd("medium","thin", at, ab))

        for i,(hall,rng) in enumerate(br_rows):
            r   = row+i
            top = bft  if i==0  else ("thin" if bft else None)
            bot = blb  if i==span-1  else "thin"
            sw(ws,r,2,rng,       F28,AL, brd("thin","thin",  top,bot))
            sw(ws,r,3,int(hall), F28,AC, brd("thin","medium",top,bot))
        row += span

# =============================================================================
# Hall (absentee) sheet — exact fixed layout
#
# Verified from original:
#  Row 1:    A1:J1    Institution
#  Row 2:    A2:J2    Absentee header
#  Row 4-5:  A4:A5 / B4:C5 / G4:H5 / I4:J5  (date + hall)
#  Row 7:    A7: brd(med,thin,None,thin)  B7-E7: brd(None,None,None,thin)
#            F7: brd(None,thin,None,thin)  G7:J7: "REGISTER NUMBER…"
#  --- per-dept blocks (5 fixed sections) ---
#  CE  : hdr row 8,  data A11:F12  right G11:J12 brd(thin,med,None,None)
#  CSE : hdr row 15, data A19:F21  right G19:J21 + G22:J22
#  ECE : hdr row 25, data A28:F29  right G28:J29 + G29:J29
#  ERE : hdr row 32, data A35:F36  right G35:J36 brd(thin,med,None,None)
#  IT  : hdr row 39, data A42:F43  right G42:J43 brd(thin,med,None,None)
#  --- footer ---
#  Row 46-50 (question papers / answer books / invigilator)
# =============================================================================

# (dept, hdr_row, data_row, data_rows,
#  right_brd_left, right_brd_right,   <- first row of data-right block
#  extra_right_rows)                   <- additional right-side merges
_HL = [
    ("CE",  8,  11, 2,  "thin",  "medium", []),
    ("CSE", 15, 19, 3,  None,    "medium", [(22,22)]),
    ("ECE", 25, 28, 2,  None,    "medium", [(29,29)]),
    ("ERE", 32, 35, 2,  "thin",  "medium", []),
    ("IT",  39, 42, 2,  "thin",  "medium", []),
]

def build_hall_sheet(wb, room, meta, hall_map):
    hall_no = hall_map.get(room["room_name"], room["room_name"])
    ws      = wb.create_sheet(hall_no)

    # Exact column widths
    for col,w in zip("ABCDEFGHIJ",
                     [8.71,7.57,9.43,8.71,8.71,15.57,7.71,8.71,3.14,8.71]):
        ws.column_dimensions[col].width = w

    # Exact row heights from original
    _RH = {
        1:15,2:15,3:15,4:15,5:15,6:15,7:15,8:15,9:15,10:15,
        11:15,12:15,13:15,14:15,15:15,16:15,17:15,18:15,19:15,20:15,
        21:7.5,22:15,23:15,24:15,25:15,26:15,27:15,28:15,29:9.75,
        30:15,31:15,32:15,33:15,34:15,35:15,36:15,37:15,38:15,39:15,
        40:15,41:15,42:15,43:10.5,44:15,45:15,
        46:24.95,47:20.1,48:20.1,50:15.75,
    }
    for r,h in _RH.items():
        ws.row_dimensions[r].height = h

    F12b=fnt("Cambria",12,True);  F12=fnt("Cambria",12,False)
    F11b=fnt("Cambria",11,True);  F11=fnt("Cambria",11,False)
    F9  =fnt("Cambria", 9,False)
    AC  =aln("center","center");  AL=aln("left","center")

    # ── Row 1: Institution ────────────────────────────────────────────────
    mw(ws,1,1,1,10, INSTITUTION,
       F12b, AC, brd("medium","medium","medium",None))

    # ── Row 2: Absentee header ────────────────────────────────────────────
    mw(ws,2,1,2,10, ABSENTEE_HDR,
       F11b, AC, brd("medium","medium",None,None))

    # ── Rows 4-5: Date / Hall No ──────────────────────────────────────────
    mw(ws,4,1,5,1,  "DATE : ",  F12,AC, brd("medium","thin","thin","thin"))
    mw(ws,4,2,5,3,  EXAM_DATE,  F12,AC, brd("thin","thin","thin","thin"))
    mw(ws,4,7,5,8,  "Hall No.", F12,AC, brd("thin","thin","thin","thin"))
    mw(ws,4,9,5,10, hall_no,    F12,AC, brd("thin","medium","thin","thin"))
    # Non-merged gap cells D-F in rows 4-5 need explicit Cambria 11 font
    for rr in [4,5]:
        for cc in [4,5,6]:
            sw(ws,rr,cc, None, F11, AC, B_NONE)

    # ── Row 7: left blank merge + right label ─────────────────────────────
    # A7:F7 merged blank - set A7 (top-left of merge) with correct border
    mw(ws,7,1,7,6, None, F11, AC, brd("medium","thin",None,"thin"))
    # G7:J7: "REGISTER NUMBER OF ABSENTEES"
    mw(ws,7,7,7,10, "REGISTER NUMBER OF ABSENTEES",
       F9, AC, brd("thin","thin",None,"thin"))

    # ── Per-dept sections ─────────────────────────────────────────────────
    bd = room_students(room)
    sm = subj_map(room, meta)

    for (dept, hdr_row, data_row, data_nrows,
         r_brd_L, r_brd_R, extra_rights) in _HL:

        # Header row: dept (A, single row), subject (B:F), dept-right (G:J)
        sw(ws, hdr_row, 1, dept,
           F12b, aln("center",None), brd("medium","thin","thin","thin"))
        mw(ws, hdr_row,2, hdr_row,6,
           "Subject: "+sm.get(dept,""), F12, AL, brd("thin","thin","thin","thin"))
        mw(ws, hdr_row,7, hdr_row,10,
           dept, F12b, aln("center",None), brd("thin","thin","thin","thin"))

        # Data area left (A:F merged, data_nrows tall)
        mw(ws, data_row,1, data_row+data_nrows-1,6,
           None, F12, aln("center","center",True),
           brd("medium","thin",None,None))

        # Right-side: individual cells G:J per data row (matching original exactly)
        # CE/CSE/ECE use bold=True; ERE/IT use bold=False (from original)
        r_fnt = F12  if dept == "IT" else F12b
        for rr in range(data_row, data_row+data_nrows):
            sw(ws,rr,7,  None, r_fnt, None, brd(r_brd_L, None,    None, None))
            sw(ws,rr,8,  None, r_fnt, None, brd(None,    None,    None, None))
            sw(ws,rr,9,  None, r_fnt, None, brd(None,    None,    None, None))
            sw(ws,rr,10, None, r_fnt, None, brd(None,    r_brd_R, None, None))

        # Extra right blank merges (G22:J22, G29:J29 in original)
        for (er1,er2) in extra_rights:
            mw(ws,er1,7,er2,10, None, F12b, None, brd(None,"medium",None,None))

        # Fill data value
        data = bd.get(dept)
        if data and (data["r"] or data["s"]):
            parts = []
            if data["r"]: parts.append(make_range_str(data["r"]))
            if data["s"]: parts.append(make_range_str(data["s"]))
            val  = "\n".join(parts)
            wrap = len(parts)>1
        else:
            val  = "NIL"
            wrap = False
        c = ws.cell(data_row,1)
        c.value=val; c.font=F12
        c.alignment=aln("center","center",wrap)
        c.border=brd("medium","thin",None,None)

    # ── Footer rows 46-50 ─────────────────────────────────────────────────
    # Row 46
    mw(ws,46,1,46,3,"No. of Question Papers received  ",
       F11,AC,brd("medium","thin","thin","thin"))
    sw(ws,46,4,None,F11,AC,brd("thin",None,"thin","thin"))    # D46
    mw(ws,46,5,46,6,"Issued ",F11,AC, brd("thin","thin","thin","thin"))
    sw(ws,46,7,None,F11,AC,brd("thin","thin","thin","thin"))   # G46
    mw(ws,46,8,46,9,"Balance",F11,AC,brd("thin","thin","thin","thin"))
    sw(ws,46,10,None,F11,AC,brd("thin","medium","thin","thin")) # J46

    # Row 47
    mw(ws,47,1,47,3,"No. of Answer Books received ",
       F11,AC,brd("medium","thin","thin","thin"))
    sw(ws,47,4,None,F11,AC,brd("thin",None,"thin","thin"))
    mw(ws,47,5,47,6,"Issued ",F11,AC,brd("thin","thin","thin","thin"))
    sw(ws,47,7,None,F11,AC,brd("thin","thin","thin","thin"))
    mw(ws,47,8,47,9,"Balance",F11,AC,brd("thin","thin","thin","thin"))
    sw(ws,47,10,None,F11,AC,brd("thin","medium","thin","thin"))

    # Row 48
    mw(ws,48,1,48,4,"Invigilator's Name & Designation",
       F11b,AC,brd("medium","thin","thin","thin"))
    mw(ws,48,5,48,7,"Department",     F11b,AC,brd("thin","thin","thin","thin"))
    mw(ws,48,8,48,10,"Dated Signature",F11b,AC,brd("thin","thin","thin","thin"))

    # Rows 49-50: signature blanks
    mw(ws,49,1,50,4, None,F11,AC,brd("medium","thin","thin","thin"))
    mw(ws,49,5,50,7, None,F11,AC,brd("thin","thin","thin","thin"))
    mw(ws,49,8,50,10,None,F11,AC,brd("thin","thin","thin","thin"))

# =============================================================================
# DOOR sheet — pixel-perfect match to original
#
# Col A: Calibri 11, left=medium border every row (print-area left edge)
# Col A row 1:  also top=medium
# Col A row 18: also bottom=medium (end of first room block)
# Grid cells (B,C,E,G,H): Calibri 14 bold, B_ALL border
# Separator cols D,F in grid: Calibri 14 bold, no border
# Summary rows: Cambria 11
# =============================================================================

_DOOR_COLS = [2, 3, 5, 7, 8]   # B, C, E, G, H

def _door_col_a(ws, r, extra_top=False, extra_bot=False):
    """Set col A style for DOOR sheet (medium-left border on every row)."""
    f  = fnt("Calibri",11,False)
    t  = "medium" if extra_top  else None
    b  = "medium" if extra_bot  else None
    sw(ws, r, 1, None, f, None, brd("medium",None,t,b))

def build_door_sheet(wb, rooms, hall_map):
    ws = wb.create_sheet("DOOR")

    for col,w in zip("ABCDEFGHIJ",
                     [1.14,17.29,18.0,2.43,18.0,2.14,19.14,17.0,1.57,2.14]):
        ws.column_dimensions[col].width = w

    F14b = fnt("Calibri",14,True)
    F18  = fnt("Calibri",18,False)
    F11b = fnt("Cambria",11,True)
    F11  = fnt("Cambria",11,False)
    AC   = aln("center","center")
    AL   = aln("left","center")
    ALT  = aln("left","top",True)

    cur = 1   # current row in the worksheet

    for room in rooms:
        hall_no   = hall_map.get(room["room_name"], room["room_name"])
        rows      = room["rows"]
        block_start = cur

        # ── 3-row title header ────────────────────────────────────────────
        # Row +0: "APJ ABDUL KALAM…"  B:H merged, top-medium border
        _door_col_a(ws, cur, extra_top=(cur==block_start))
        mw(ws,cur,2,cur,8,
           "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY",
           F14b, AC, brd(None,None,"medium",None))
        ws.row_dimensions[cur].height = 18.75; cur += 1

        # Row +1: exam line
        _door_col_a(ws, cur)
        mw(ws,cur,2,cur,8, DOOR_EXAM, F14b, AC, B_NONE)
        ws.row_dimensions[cur].height = 18.75; cur += 1

        # Row +2: "SEATING PLAN"
        _door_col_a(ws, cur)
        mw(ws,cur,3,cur,7, "SEATING PLAN", F14b, AC, B_NONE)
        # H col of row+2 needs Calibri 14 bold (original)
        sw(ws,cur,8, None, F14b, None, B_NONE)
        ws.row_dimensions[cur].height = 18.75; cur += 1

        # Row +3: date (B:C) | hall (G:H)
        _door_col_a(ws, cur)
        mw(ws,cur,2,cur,3, EXAM_DATE+" ", F18, AL, B_ALL)
        mw(ws,cur,7,cur,8, f"HALL : {hall_no}", F18, AL, B_ALL)
        ws.row_dimensions[cur].height = 23.25; cur += 1

        # Row +4: blank (default height)
        _door_col_a(ws, cur)
        cur += 1

        # Row +5: column number labels 1-5 in B,C,E,G,H
        # Separator cols D,F also have Calibri 14 bold (no border)
        _door_col_a(ws, cur)
        for idx,cp in enumerate(_DOOR_COLS, 1):
            sw(ws,cur,cp, str(idx), F14b, AC, B_ALL)
        sw(ws,cur,4, None, fnt("Calibri",11), None, B_NONE)   # D: separator
        sw(ws,cur,6, None, fnt("Calibri",11), None, B_NONE)   # F: separator
        ws.row_dimensions[cur].height = 18.75; cur += 1

        # ── Seating grid ──────────────────────────────────────────────────
        max_rn   = max((rn for rn,_ in rows), default=0)
        col_data = defaultdict(dict)
        for rn,vals in rows:
            for ci,v in enumerate(vals):
                if v: col_data[ci][rn] = v

        for rnum in range(1, max_rn+1):
            if not any(col_data[ci].get(rnum) for ci in range(5)): continue
            _door_col_a(ws, cur)
            for ci,cp in enumerate(_DOOR_COLS):
                v = col_data[ci].get(rnum)
                if v:
                    sw(ws,cur,cp, v, F14b, AC, B_ALL)
            # Separator cols D,F always get Calibri 14 bold in grid rows
            sw(ws,cur,4, None, F14b, None, B_NONE)
            sw(ws,cur,6, None, F14b, None, B_NONE)
            ws.row_dimensions[cur].height = 18.75
            cur += 1

        # ── Blank row after grid (col A continues; row bottom = None) ─────
        _door_col_a(ws, cur)
        cur += 1

        # ── Branch summary ────────────────────────────────────────────────
        bd     = room_students(room)
        reg_bd = {br:data["r"] for br,data in bd.items() if data["r"]}
        sup_bd = {br:data["s"] for br,data in bd.items() if data["s"]}

        for br in ordered_br(bd):
            if br not in reg_bd: continue
            _door_col_a(ws, cur)
            sw(ws,cur,2, br,  F11b, AL, B_ALL)
            mw(ws,cur,3,cur,8, make_range_str(reg_bd[br]), F11, AL, B_ALL)
            ws.row_dimensions[cur].height = 20.1; cur += 1

        if sup_bd:
            parts = [f"{br} - {make_range_str(sup_bd[br])}"
                     for br in ordered_br(bd) if br in sup_bd]
            # "Supply" label: B merged 2 rows, C:H merged 2 rows (original B16:B17, C16:H17)
            _door_col_a(ws, cur)
            _door_col_a(ws, cur+1)
            mw(ws,cur,  2,cur+1,2, "Supply",          F11b, AL,  B_ALL)
            mw(ws,cur,  3,cur+1,8, ", ".join(parts),  F11,  ALT, B_ALL)
            ws.row_dimensions[cur].height   = 14.25
            ws.row_dimensions[cur+1].height = 32.25
            cur += 2

        # ── Gap rows between rooms ────────────────────────────────────────
        # Col A continues left-medium border; last gap row gets bottom=medium
        gap_heights = [17.25, 17.25, 17.25, 17.25, 15.75]
        for gi, gh in enumerate(gap_heights):
            is_last_gap = (gi == len(gap_heights)-1)
            _door_col_a(ws, cur, extra_bot=is_last_gap)
            # Rows 18 equiv: all B:H cols get bottom=medium on last gap row
            if is_last_gap:
                for c in range(2,9):
                    sw(ws,cur,c, None, fnt("Calibri",11), None,
                       brd(None,None,None,"medium"))
            ws.row_dimensions[cur].height = gh; cur += 1

# =============================================================================
# Entry point
# =============================================================================

def main():
    print(f"Parsing: {TXT_FILE}")
    meta, rooms = parse_txt(TXT_FILE)
    if not rooms:
        print("ERROR: No rooms found.", file=sys.stderr); sys.exit(1)

    print(f"Found {len(rooms)} rooms, {len(meta['subjects'])} subjects\n")
    for room in rooms:
        hall = ROOM_TO_HALL.get(room["room_name"],"?")
        bd   = room_students(room)
        tr   = sum(len(d["r"]) for d in bd.values())
        ts   = sum(len(d["s"]) for d in bd.values())
        print(f"  {room['room_name']:8s} -> Hall {hall}: "
              f"{tr} regular + {ts} supply = {tr+ts} total")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nBuilding MASTER...")
    build_master(wb, meta, rooms, ROOM_TO_HALL)

    print("Building hall sheets...")
    for room in rooms:
        hall = ROOM_TO_HALL.get(room["room_name"], room["room_name"])
        print(f"  Hall {hall}")
        build_hall_sheet(wb, room, meta, ROOM_TO_HALL)

    print("Building DOOR sheet...")
    build_door_sheet(wb, rooms, ROOM_TO_HALL)

    # Save — auto-fallback if file is open in Excel
    save_path = OUTPUT_EXCEL
    try:
        wb.save(save_path)
    except PermissionError:
        base, ext = os.path.splitext(OUTPUT_EXCEL)
        for i in range(1, 21):
            save_path = f"{base}_{i}{ext}"
            try:
                wb.save(save_path)
                print(f"\nNOTE: '{OUTPUT_EXCEL}' is open — close it to overwrite directly.")
                print(f"      Saved to '{save_path}' instead.")
                break
            except PermissionError:
                continue
        else:
            print("ERROR: Cannot save. Close the output file and retry."); sys.exit(1)

    print(f"\nDone!  Saved: {save_path}")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
